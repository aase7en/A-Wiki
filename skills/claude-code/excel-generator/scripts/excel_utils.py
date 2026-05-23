"""
excel_utils.py — Wiki Excel Generator Utilities
ใช้ patterns จาก excel-generator skill (SKILL.md)

Usage:
    from excel_utils import ExcelBuilder, THEMES, HIGHLIGHT
    wb = ExcelBuilder(theme='elegant_black')
    ws = wb.new_sheet("Overview")
    wb.title(ws, "My Report", row=2)
    wb.section_header(ws, "DATA", row=5, col_span=(2, 8))
    wb.table(ws, headers, rows, start_row=6, start_col=2)
    wb.save("/tmp/output.xlsx")
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
from openpyxl.chart import BarChart, LineChart, Reference
from datetime import date


# ── Theme Definitions ──────────────────────────────────────────────────────────

THEMES = {
    'elegant_black': {
        'primary': '2D2D2D', 'light': 'E5E5E5',
        'chart_colors': ['2D2D2D', '4A4A4A', '6B6B6B', '8C8C8C'],
    },
    'corporate_blue': {
        'primary': '1F4E79', 'light': 'D6E3F0',
        'chart_colors': ['1F4E79', '2E75B6', '5B9BD5', '9DC3E6'],
    },
    'forest_green': {
        'primary': '2E5A4C', 'light': 'D4E5DE',
        'chart_colors': ['2E5A4C', '3D7A65', '56A68A', '7BBFAE'],
    },
    'teal': {
        'primary': '1A5F5F', 'light': 'D3E5E5',
        'chart_colors': ['1A5F5F', '267A7A', '3D9999', '6FBBBB'],
    },
}

# ── Semantic Colors ────────────────────────────────────────────────────────────
SEMANTIC = {
    'positive': '2E7D32',   # green — growth, success
    'negative': 'C62828',   # red — decline, loss
    'warning':  'F57C00',   # orange — caution
}

HIGHLIGHT = {
    'emphasis': 'E6F3FF',   # blue tint — important data
    'section':  'FFF3E0',   # amber tint — section dividers
    'input':    'FFFDE7',   # yellow tint — editable cells
    'success':  'E8F5E9',   # green tint — passed
    'warning':  'FFCCBC',   # red tint — needs attention
}

SERIF = 'Georgia'       # fallback for Source Serif Pro
SANS  = 'Calibri'       # fallback for Source Sans Pro


# ── Low-level Helpers ──────────────────────────────────────────────────────────

def _fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

def _side(style='thin', color='D1D1D1') -> Side:
    return Side(style=style, color=color)

def _no_side() -> Side:
    return Side(style=None)


def apply_data_block_borders(ws, r1: int, r2: int, c1: int, c2: int,
                              has_header: bool = True, theme_primary: str = '2D2D2D'):
    """Apply outer frame + horizontal inner borders to a Data Block."""
    outer   = _side('thin', 'AAAAAA')
    inner   = _side('thin', 'E0E0E0')
    hdr_bot = _side('medium', theme_primary)
    no      = _no_side()

    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            left   = outer if c == c1 else no
            right  = outer if c == c2 else no
            top    = outer if r == r1 else inner
            if has_header and r == r1:
                bottom = hdr_bot
            elif r == r2:
                bottom = outer
            else:
                bottom = inner
            ws.cell(r, c).border = Border(left=left, right=right, top=top, bottom=bottom)


# ── ExcelBuilder Class ─────────────────────────────────────────────────────────

class ExcelBuilder:
    """
    Wrapper ที่ใช้ excel-generator skill patterns ในการสร้าง Excel ไฟล์

    Example:
        eb = ExcelBuilder('corporate_blue')
        ws = eb.new_sheet("Report")
        eb.title(ws, "Drug Order Report")
        eb.section_header(ws, "ITEMS", row=5, col_span=(2, 8))
        eb.table(ws,
            headers=["#", "Name", "Qty", "Price"],
            rows=[[1, "Amoxicillin", 10, 85.0]],
            start_row=6, start_col=2,
            num_cols={3: '#,##0', 4: '#,##0.00'}
        )
        eb.save("/tmp/out.xlsx")
    """

    def __init__(self, theme: str = 'elegant_black'):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)   # remove default sheet
        self.t = THEMES.get(theme, THEMES['elegant_black'])
        self._first = True

    # ── Sheet ──────────────────────────────────────────────────────────────────

    def new_sheet(self, name: str):
        ws = self.wb.create_sheet(name)
        ws.sheet_view.showGridLines = False
        ws.column_dimensions['A'].width = 3   # left margin
        ws.row_dimensions[1].height = 8       # top margin
        return ws

    # ── Text Elements ──────────────────────────────────────────────────────────

    def title(self, ws, text: str, row: int = 2, col_start: int = 2, col_end: int = 8):
        ws.merge_cells(f'{get_column_letter(col_start)}{row}:{get_column_letter(col_end)}{row}')
        cell = ws.cell(row, col_start, text)
        cell.font = Font(name=SERIF, size=18, bold=True, color=self.t['primary'])
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 36

    def subtitle(self, ws, text: str, row: int = 3, col_start: int = 2, col_end: int = 8):
        ws.merge_cells(f'{get_column_letter(col_start)}{row}:{get_column_letter(col_end)}{row}')
        cell = ws.cell(row, col_start, text)
        cell.font = Font(name=SANS, size=10, italic=True, color='666666')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 18

    def section_header(self, ws, text: str, row: int = 5,
                       col_span: tuple = (2, 8), with_bg: bool = True):
        c1, c2 = col_span
        ws.merge_cells(f'{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}')
        cell = ws.cell(row, c1, text)
        cell.font = Font(name=SERIF, size=13, bold=True, color=self.t['primary'])
        if with_bg:
            cell.fill = _fill(self.t['light'])
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row].height = 24

    def insights(self, ws, bullet_list: list, start_row: int,
                 col_span: tuple = (2, 8)):
        c1, c2 = col_span
        header_row = start_row
        ws.merge_cells(f'{get_column_letter(c1)}{header_row}:{get_column_letter(c2)}{header_row}')
        hcell = ws.cell(header_row, c1, "KEY INSIGHTS")
        hcell.font = Font(name=SERIF, size=13, bold=True, color=self.t['primary'])
        hcell.fill = _fill(self.t['light'])
        hcell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[header_row].height = 24

        for i, text in enumerate(bullet_list):
            r = start_row + 1 + i
            ws.merge_cells(f'{get_column_letter(c1)}{r}:{get_column_letter(c2)}{r}')
            cell = ws.cell(r, c1, f"• {text}")
            cell.font = Font(name=SANS, size=11, color='333333')
            cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[r].height = 18

        return start_row + 1 + len(bullet_list)  # next available row

    def footer(self, ws, row: int, col_span: tuple = (2, 8)):
        c1, c2 = col_span
        ws.merge_cells(f'{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}')
        cell = ws.cell(row, c1,
            f"Generated by Claude Code — {date.today().isoformat()}")
        cell.font = Font(name=SANS, size=9, italic=True, color='AAAAAA')
        cell.alignment = Alignment(horizontal='left', vertical='center')

    # ── Table ──────────────────────────────────────────────────────────────────

    def table(self, ws, headers: list, rows: list,
              start_row: int = 6, start_col: int = 2,
              col_widths: list | None = None,
              num_cols: dict | None = None,
              highlight_rows: dict | None = None) -> int:
        """
        Render a complete Data Block table.

        Args:
            headers     : list of column header strings
            rows        : list of row value lists
            start_row   : row index for header
            start_col   : column index for first column
            col_widths  : list of widths per column (optional)
            num_cols    : dict {col_offset: format_str} e.g. {3: '#,##0', 4: '#,##0.00'}
            highlight_rows: dict {row_index(0-based): HIGHLIGHT_KEY}

        Returns:
            next available row index
        """
        nc = len(headers)
        end_col = start_col + nc - 1
        num_cols = num_cols or {}
        highlight_rows = highlight_rows or {}

        # Column widths
        if col_widths:
            for i, w in enumerate(col_widths):
                ws.column_dimensions[get_column_letter(start_col + i)].width = w

        # Header row
        hr = start_row
        for i, h in enumerate(headers):
            cell = ws.cell(hr, start_col + i, h)
            cell.font = Font(name=SERIF, size=10, bold=True, color='FFFFFF')
            cell.fill = _fill(self.t['primary'])
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[hr].height = 28

        # Data rows
        for ri, row_data in enumerate(rows):
            r = start_row + 1 + ri
            row_bg = HIGHLIGHT.get(highlight_rows.get(ri, ''), None)
            bg = row_bg or ('F7F7F7' if ri % 2 == 0 else 'FFFFFF')

            for ci, val in enumerate(row_data):
                col_idx = start_col + ci
                cell = ws.cell(r, col_idx, val)
                cell.fill = _fill(bg)
                cell.font = Font(name=SANS, size=11, color='2D2D2D')

                # Number format
                if (ci + 1) in num_cols:
                    cell.number_format = num_cols[ci + 1]
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif isinstance(val, str) and len(val) > 20:
                    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            ws.row_dimensions[r].height = 20

        end_row = start_row + len(rows)
        apply_data_block_borders(ws, hr, end_row, start_col, end_col,
                                  has_header=True, theme_primary=self.t['primary'])

        return end_row + 1   # next available row

    def data_bar(self, ws, range_str: str):
        """Add data bar conditional formatting to a range."""
        ws.conditional_formatting.add(
            range_str,
            DataBarRule(start_type='min', end_type='max', color=self.t['primary'])
        )

    def color_scale(self, ws, range_str: str):
        """White → primary color scale."""
        ws.conditional_formatting.add(
            range_str,
            ColorScaleRule(start_type='min', start_color='FFFFFF',
                           end_type='max',   end_color=self.t['primary'])
        )

    # ── Save ───────────────────────────────────────────────────────────────────

    def save(self, path: str):
        self.wb.save(path)
        print(f"✅ Excel saved → {path}")
        return path
