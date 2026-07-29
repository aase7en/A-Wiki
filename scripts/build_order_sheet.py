#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_order_sheet.py — ใบสั่งซื้อยาร้านภูฟาร์มาซี: JSON → XLSX → (ผู้ใช้กรอกจำนวน) → ข้อความ LINE

3 โหมด
  build  : items.json  → ใบสั่งซื้อ .xlsx ตาม template (12 หมวด, ช่องกรอกสีเหลือง, สูตรรวมเงิน,
                          ชีตสรุปตามหมวด, เติม "ทุน/หน่วย" จาก medi_list.db ให้อัตโนมัติ)
  line   : .xlsx ที่กรอกจำนวนแล้ว → ข้อความกระชับสำหรับ copy วางใน LINE
  cost   : ค้นราคาทุน+หน่วย+คงเหลือ จาก medi_list.db (debug/ตรวจสอบ)

รูปแบบ items.json
[
  {"category": "1. ยาแก้ปวด / ลดไข้",
   "name": "PARA CEMOL (ชนิดเม็ดขาว กลม)",
   "strength": "Paracetamol 500 mg",
   "pack": "100 เม็ด/ขวด",
   "qty": 12, "unit": "ขวด",
   "note": "", "cost": null}
]
qty ปล่อย null/0 ได้ (ให้ผู้ใช้กรอกในชีต) — cost ถ้าไม่ใส่จะ lookup ให้เอง

Usage:
  python3 scripts/build_order_sheet.py build items.json --out order.xlsx --po PO-2026-08-01
  python3 scripts/build_order_sheet.py line  order.xlsx
  python3 scripts/build_order_sheet.py cost  "betadine" "minny 28"
"""
from __future__ import annotations

import argparse
import json
import pathlib
import re
import sqlite3
import sys
from datetime import date

ROOT = pathlib.Path(__file__).resolve().parents[1]
MEDI_DB = ROOT / "wiki/entities/pharmacy/medi_list.db"

FONT = "Tahoma"
NAVY = "1F3864"
HEADERS = ["ลำดับ", "หมวด", "ชื่อยา / สินค้า", "ตัวยา / ความแรง", "รูปแบบ / ขนาดบรรจุ",
           "จำนวนสั่ง", "หน่วย", "ราคา/หน่วย (บาท)", "รวมเงิน (บาท)", "ผู้จำหน่าย", "หมายเหตุ"]
HEADER_ROW = 5

# ---------------------------------------------------------------- cost lookup


_UNIT_ALIAS = {"cc": "ml", "gm": "g", "gram": "g", "grams": "g", "mgs": "mg",
               "tabs": "s", "tab": "s", "caps": "s", "cap": "s"}


def _tokens(text: str) -> set[str]:
    """แตกคำ + แยกตัวเลขออกจากหน่วย + ปรับหน่วยให้เป็นรูปมาตรฐาน (15ml / 15 CC → {15, ml})."""
    out: set[str] = set()
    for raw in re.split(r"[^0-9A-Za-z฀-๿]+", (text or "").lower()):
        if not raw:
            continue
        for part in re.findall(r"\d+(?:\.\d+)?|[^\W\d_]+", raw, flags=re.UNICODE):
            out.add(_UNIT_ALIAS.get(part, part))
    return out - {""}


def _fts_query(text: str) -> str:
    toks = [t for t in re.split(r"[^0-9A-Za-z฀-๿]+", text) if len(t) > 1]
    return " OR ".join(f'"{t}"' for t in toks[:6])


def lookup_cost(name: str, db: pathlib.Path = MEDI_DB) -> dict | None:
    """คืน {name, unit, cost, on_hand, score} จาก medi_list.db หรือ None ถ้าไม่มั่นใจพอ."""
    if not db.exists():
        return None
    q = _fts_query(name)
    if not q:
        return None
    conn = sqlite3.connect(db)
    try:
        rows = conn.execute(
            "SELECT m.name, m.unit, m.cost, m.on_hand, bm25(medi_fts) AS rank "
            "FROM medi_fts JOIN medi m ON m.id = medi_fts.rowid "
            "WHERE medi_fts MATCH ? ORDER BY rank LIMIT 5", (q,)).fetchall()
    except sqlite3.OperationalError:
        return None
    finally:
        conn.close()
    if not rows:
        return None

    want = _tokens(name)
    # ตัวเลขขนาด/ความแรงในคำค้น — ถ้ามีแล้วผู้สมัครไม่มีเลย = คนละขนาด ห้ามจับคู่
    want_size = {t for t in want if any(ch.isdigit() for ch in t)}
    want_word = want - want_size
    # token ที่ยาวที่สุด = ชื่อการค้าที่จำเพาะที่สุด — ต้องเจอในผลลัพธ์เสมอ
    key_word = max(want_word, key=len) if want_word else ""

    best, best_s = None, 0.0
    for nm, unit, cost, on_hand, _rank in rows:
        got = _tokens(nm)
        if want_size and not (want_size & got):
            continue                       # กัน "MYDA-B 15g" → "Myda-B 25g"
        if key_word and key_word not in got:
            continue                       # กัน "betadine 15ml"→"NASOL 15ml", "Nizoral cream"→"ZEMA Cream"
        s = len(want & got) / max(len(want), 1)
        if s > best_s:
            best, best_s = (nm, unit, cost, on_hand), s
    if not best or best_s < 0.5:          # ต่ำกว่านี้ = เดา — ปล่อยว่างดีกว่าใส่ผิด
        return None
    return {"name": best[0], "unit": best[1], "cost": best[2],
            "on_hand": best[3], "score": round(best_s, 2)}


# ---------------------------------------------------------------- build xlsx


def build(items: list[dict], out: pathlib.Path, po: str, order_date: str, buyer: str) -> dict:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor=NAVY)
    CAT_FILL = PatternFill("solid", fgColor="D9E2F3")
    INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
    TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
    thin = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    BLUE = Font(name=FONT, size=10, color="0000FF")

    wb = Workbook()
    ws = wb.active
    ws.title = "ใบสั่งซื้อ"

    ws["A1"] = "ใบสั่งซื้อยาและเวชภัณฑ์ — ร้านยาภูฟาร์มาซี"
    ws["A1"].font = Font(name=FONT, size=15, bold=True, color=NAVY)
    ws.merge_cells("A1:K1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws["A2"], ws["B2"] = "วันที่สั่งซื้อ:", order_date
    ws["D2"], ws["E2"] = "ผู้สั่งซื้อ:", buyer
    ws["H2"], ws["I2"] = "เลขที่ใบสั่งซื้อ:", po
    for c in ("B2", "E2", "I2"):
        ws[c].font = BLUE
    for c in ("A2", "D2", "H2"):
        ws[c].font = Font(name=FONT, size=10, bold=True)

    ws["A3"] = ("วิธีใช้: กรอกเฉพาะช่องพื้นสีเหลือง (จำนวนสั่ง / ราคาต่อหน่วย / ผู้จำหน่าย) "
                "— คอลัมน์ 'รวมเงิน' และยอดรวมคำนวณอัตโนมัติ | ราคาที่เติมไว้คือ 'ทุนล่าสุด' จากสต๊อกร้าน")
    ws["A3"].font = Font(name=FONT, size=9, italic=True, color="808080")
    ws.merge_cells("A3:K3")

    for col, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=HEADER_ROW, column=col, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[HEADER_ROW].height = 32

    row, seq, cur, drows = HEADER_ROW + 1, 0, None, []
    filled_cost = 0

    for it in items:
        cat = it.get("category", "อื่น ๆ")
        if cat != cur:
            cur = cat
            ws.cell(row=row, column=1, value=cat).font = Font(name=FONT, size=10, bold=True, color=NAVY)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
            for col in range(1, 12):
                ws.cell(row=row, column=col).fill = CAT_FILL
                ws.cell(row=row, column=col).border = BORDER
            row += 1

        cost = it.get("cost")
        if cost is None:
            hit = lookup_cost(it.get("name", ""))
            if hit and hit.get("cost"):
                cost = round(float(hit["cost"]), 2)
                filled_cost += 1

        seq += 1
        ws.cell(row=row, column=1, value=seq)
        ws.cell(row=row, column=2, value=cat)
        ws.cell(row=row, column=3, value=it.get("name", ""))
        ws.cell(row=row, column=4, value=it.get("strength", ""))
        ws.cell(row=row, column=5, value=it.get("pack", ""))
        ws.cell(row=row, column=6, value=it.get("qty") or None)
        ws.cell(row=row, column=7, value=it.get("unit", ""))
        ws.cell(row=row, column=8, value=cost)
        ws.cell(row=row, column=9, value=f'=IF(OR(F{row}="",H{row}=""),0,F{row}*H{row})')
        ws.cell(row=row, column=11, value=it.get("note", ""))

        for col in range(1, 12):
            c = ws.cell(row=row, column=col)
            c.font = Font(name=FONT, size=10)
            c.border = BORDER
            c.alignment = Alignment(vertical="center",
                                    horizontal="center" if col in (1, 6, 7, 8, 9) else "left",
                                    wrap_text=col in (3, 4, 5, 11))
        for col in (6, 8, 10):
            ws.cell(row=row, column=col).fill = INPUT_FILL
            ws.cell(row=row, column=col).font = BLUE
        ws.cell(row=row, column=6).number_format = "#,##0"
        ws.cell(row=row, column=8).number_format = "#,##0.00;(#,##0.00);-"
        ws.cell(row=row, column=9).number_format = "#,##0.00;(#,##0.00);-"
        ws.cell(row=row, column=2).font = Font(name=FONT, size=9, color="808080")
        ws.cell(row=row, column=11).font = Font(name=FONT, size=9,
                                                color="C00000" if it.get("note") else "000000")
        drows.append(row)
        row += 1

    first_d, last_d = drows[0], drows[-1]
    trow = row + 1
    ws.cell(row=trow, column=3, value="รวมทั้งสิ้น")
    ws.cell(row=trow, column=6, value=f"=SUM(F{first_d}:F{last_d})")
    ws.cell(row=trow, column=9, value=f"=SUM(I{first_d}:I{last_d})")
    for col in range(1, 12):
        c = ws.cell(row=trow, column=col)
        c.fill = TOTAL_FILL
        c.font = Font(name=FONT, size=11, bold=True, color=NAVY)
        c.border = BORDER
        c.alignment = Alignment(horizontal="right" if col == 3 else "center", vertical="center")
    ws.cell(row=trow, column=6).number_format = "#,##0"
    ws.cell(row=trow, column=9).number_format = "#,##0.00;(#,##0.00);-"

    for i, w in enumerate([6, 24, 38, 30, 22, 10, 9, 14, 14, 18, 30], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.column_dimensions["B"].hidden = True
    ws.freeze_panes = f"A{HEADER_ROW + 1}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:K{last_d}"
    ws.sheet_view.showGridLines = False

    # ---- ชีตสรุป ----
    ws2 = wb.create_sheet("สรุปตามหมวด")
    ws2["A1"] = "สรุปยอดสั่งซื้อแยกตามหมวดการรักษา"
    ws2["A1"].font = Font(name=FONT, size=13, bold=True, color=NAVY)
    ws2.merge_cells("A1:D1")
    for col, h in enumerate(["หมวดการรักษา", "จำนวนรายการ", "จำนวนหน่วยที่สั่ง", "รวมเงิน (บาท)"], start=1):
        c = ws2.cell(row=3, column=col, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    cats: list[str] = []
    for it in items:
        c = it.get("category", "อื่น ๆ")
        if c not in cats:
            cats.append(c)

    r2 = 4
    for cat in cats:
        ws2.cell(row=r2, column=1, value=cat)
        ws2.cell(row=r2, column=2, value=f"=COUNTIF('ใบสั่งซื้อ'!$B${first_d}:$B${last_d},$A{r2})")
        ws2.cell(row=r2, column=3, value=f"=SUMIF('ใบสั่งซื้อ'!$B${first_d}:$B${last_d},$A{r2},'ใบสั่งซื้อ'!$F${first_d}:$F${last_d})")
        ws2.cell(row=r2, column=4, value=f"=SUMIF('ใบสั่งซื้อ'!$B${first_d}:$B${last_d},$A{r2},'ใบสั่งซื้อ'!$I${first_d}:$I${last_d})")
        for col in range(1, 5):
            c = ws2.cell(row=r2, column=col)
            c.font = Font(name=FONT, size=10)
            c.border = BORDER
            c.alignment = Alignment(horizontal="left" if col == 1 else "center", vertical="center")
        ws2.cell(row=r2, column=3).number_format = "#,##0"
        ws2.cell(row=r2, column=4).number_format = "#,##0.00;(#,##0.00);-"
        r2 += 1

    ws2.cell(row=r2, column=1, value="รวมทั้งสิ้น")
    for col, letter in ((2, "B"), (3, "C"), (4, "D")):
        ws2.cell(row=r2, column=col, value=f"=SUM({letter}4:{letter}{r2-1})")
    for col in range(1, 5):
        c = ws2.cell(row=r2, column=col)
        c.font = Font(name=FONT, size=11, bold=True, color=NAVY)
        c.fill = TOTAL_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="left" if col == 1 else "center", vertical="center")
    ws2.cell(row=r2, column=3).number_format = "#,##0"
    ws2.cell(row=r2, column=4).number_format = "#,##0.00;(#,##0.00);-"
    for col, w in zip("ABCD", [34, 16, 20, 18]):
        ws2.column_dimensions[col].width = w
    ws2.sheet_view.showGridLines = False

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return {"items": seq, "rows": [first_d, last_d], "cost_filled": filled_cost,
            "categories": len(cats), "out": str(out)}


# ---------------------------------------------------------------- xlsx → LINE


def to_line(xlsx: pathlib.Path, skip_zero: bool = True) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx, data_only=True)
    ws = wb["ใบสั่งซื้อ"] if "ใบสั่งซื้อ" in wb.sheetnames else wb.worksheets[0]

    po = ws["I2"].value or ""
    order_date = ws["B2"].value or ""
    if hasattr(order_date, "strftime"):
        order_date = order_date.strftime("%d/%m/%Y")

    out_lines: list[str] = []
    body: list[str] = []
    n = 0
    cur_cat = None

    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        seq, cat, name, strength, pack, qty, unit = (list(row) + [None] * 7)[:7]
        if isinstance(seq, str) or seq is None or not name:
            continue
        if str(name).strip() == "รวมทั้งสิ้น":
            break
        if skip_zero and not qty:
            continue
        if cat != cur_cat:
            cur_cat = cat
            label = re.sub(r"^\d+\.\s*", "", str(cat or "")).strip()
            body.append(f"\n■ {label}")
        n += 1
        parts = [str(name).strip()]
        if strength and str(strength).strip() not in ("", "-", "—"):
            parts.append(str(strength).strip())
        if pack and str(pack).strip() not in ("", "-", "—"):
            parts.append(str(pack).strip())
        tail = f"{int(qty)} {unit}".strip() if qty else "___"
        body.append(f"{n}. {' '.join(parts)} = {tail}")

    header = "📋 ใบสั่งยา ร้านภูฟาร์มาซี"
    if order_date:
        header += f" | {order_date}"
    if po:
        header += f" | {po}"
    out_lines.append(header)
    out_lines.append(f"รวม {n} รายการ")
    out_lines.extend(body)
    out_lines.append("\nรบกวนแจ้งราคาและกำหนดส่งด้วยครับ ขอบคุณครับ")
    return "\n".join(out_lines)


# ---------------------------------------------------------------- cli


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="cmd", required=True)

    b = sub.add_parser("build", help="items.json → ใบสั่งซื้อ .xlsx")
    b.add_argument("items")
    b.add_argument("--out", default="order.xlsx")
    b.add_argument("--po", default=f"PO-{date.today():%Y-%m}-01")
    b.add_argument("--date", default=f"{date.today():%d/%m/%Y}")
    b.add_argument("--buyer", default="ศุภศิษฎิ์ คงสุวรรณ")

    l = sub.add_parser("line", help=".xlsx ที่กรอกจำนวนแล้ว → ข้อความ LINE")
    l.add_argument("xlsx")
    l.add_argument("--all", action="store_true", help="รวมรายการที่ยังไม่กรอกจำนวนด้วย")

    c = sub.add_parser("cost", help="ค้นราคาทุนจาก medi_list.db")
    c.add_argument("names", nargs="+")

    a = ap.parse_args()

    if a.cmd == "build":
        items = json.loads(pathlib.Path(a.items).read_text(encoding="utf-8"))
        info = build(items, pathlib.Path(a.out), a.po, a.date, a.buyer)
        print(json.dumps(info, ensure_ascii=False, indent=2))
        return 0

    if a.cmd == "line":
        print(to_line(pathlib.Path(a.xlsx), skip_zero=not a.all))
        return 0

    for nm in a.names:
        hit = lookup_cost(nm)
        print(f"{nm:35s} → {json.dumps(hit, ensure_ascii=False) if hit else 'ไม่พบ (ปล่อยว่างให้ผู้ใช้กรอก)'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
