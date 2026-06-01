#!/usr/bin/env python3
"""
compare_delivery.py — เปรียบเทียบ order list กับ ใบส่งสินค้าจริง

Usage:
  python3 scripts/compare_delivery.py --delivery <delivery.json> [order_file.txt]
  python3 scripts/compare_delivery.py \\
    --delivery wiki/entities/pharmacy/deliveries/20260520_SCH_ORNC2605P010859.json \\
    --export excel

Input (stdin or file): รายการที่สั่ง 1 บรรทัด/รายการ (รองรับ "name: qty" หรือ "name qty")
"""
import json, re, sys, pathlib, argparse
from datetime import datetime
from difflib import SequenceMatcher

ROOT       = pathlib.Path(__file__).parent.parent
EXPORTS    = ROOT / "wiki/entities/pharmacy/exports"

# ── Re-use normalisation from pharmacy_lookup ────────────────────────────────
def norm(s: str) -> str:
    return re.sub(r"[\s.\-_()/\[\]]+", " ", s.lower()).strip()

_QTY_COLON = re.compile(r":\s*(\d+)\s*$")
_QTY_TIMES = re.compile(r"\s*[×xX]\s*(\d+)\s*$")

def parse_item(line: str) -> tuple[str, int]:
    m = _QTY_COLON.search(line) or _QTY_TIMES.search(line)
    if m:
        return line[:m.start()].strip(), int(m.group(1))
    # Try last token as qty: "BETADINE 15ml ขวด 12"
    parts = line.rsplit(None, 1)
    if len(parts) == 2 and parts[1].isdigit():
        return parts[0].strip(), int(parts[1])
    return line.strip(), 0

def _GENERIC_WORDS():
    return {'cream','tablet','tablets','capsule','capsules','syrup','solution',
            'injection','drop','drops','ointment','lotion','gel','spray','powder',
            'suspension','forte','plus','oral','topical','patch'}

def token_score(a: str, b: str) -> float:
    na, nb = norm(a), norm(b)
    if na == nb: return 1.0
    if len(na) > 3 and na in nb: return 0.92
    if len(nb) > 3 and nb in na: return 0.88
    ta, tb = set(na.split()), set(nb.split())
    overlap = len(ta & tb)
    if overlap == 0:
        return SequenceMatcher(None, na, nb).ratio()
    jaccard = overlap / len(ta | tb)
    if ta and ta <= tb: return max(jaccard, 0.85)
    base = max(jaccard, SequenceMatcher(None, na, nb).ratio())
    gw = _GENERIC_WORDS()
    distinctive = {t for t in (ta & tb) if len(t) >= 4 and not t[:1].isdigit() and t not in gw}
    if distinctive:
        return max(base, 0.75)
    return base


# ── Dedup (same as pharmacy_lookup) ─────────────────────────────────────────

def dedup_items(raw_lines: list[str]) -> tuple[list[tuple[str,int]], list[dict]]:
    parsed = []
    for line in raw_lines:
        if not line.strip(): continue
        drug, qty = parse_item(line)
        parsed.append({"drug": drug, "qty": qty, "key": norm(drug)})

    groups: dict[str, list] = {}
    key_order: list[str] = []
    for p in parsed:
        k = p["key"]
        if k not in groups: key_order.append(k)
        groups.setdefault(k, []).append(p)

    deduped: list[tuple[str,int]] = []
    merges: list[dict] = []
    for key in key_order:
        items = groups[key]
        total_qty = sum(i["qty"] for i in items)
        name = items[0]["drug"]
        deduped.append((name, total_qty))
        if len(items) > 1:
            merges.append({"drug": name, "count": len(items),
                           "orig_qtys": [i["qty"] for i in items],
                           "merged_qty": total_qty})
    return deduped, merges


# ── SCH-specific name normalization ─────────────────────────────────────────
# Maps order-side patterns → SCH naming patterns so fuzzy matching works better

_SCH_SUBS = [
    # Normal saline — map to NSS with volume unit normalised
    (r'klean\s*&?\s*kare\s+normal\s+saline\s+100\s*ml', 'n.s.s 100cc'),
    (r'klean\s*&?\s*kare\s+normal\s+saline\s+500\s*ml', 'n.s.s 500cc'),
    (r'klean\s*&?\s*kare\s+normal\s+saline\s+1000\s*ml', 'n.s.s 1000cc'),
    (r'klean\s*&?\s*kare\s+normal\s+saline', 'n.s.s'),
    (r'normal\s+saline\s+(\d+)\s*ml', r'n.s.s \1cc'),
    (r'normal\s+saline', 'n.s.s'),
    # Bandage
    (r'ผ้าก๊อตยืดหยุ่น', 'ผ้าก๊อต ลองเมต ชนิดยืดหยุ่น'),
    # Tiger balm plasters
    (r'พลาสเตอร์ตราเสือ.*?สูตรร้อน.*?10x14', 'พลาสเตอร์ ปิดปวดเลือดแดงใหญ่ 10x14'),
    (r'พลาสเตอร์ตราเสือ.*?สูตรร้อน.*?10x7', 'พลาสเตอร์ ปิดปวดเลือดแดง เล็ก 7x10'),
    (r'พลาสเตอร์ตราเสือ.*?สูตรเย็น.*?10x14', 'พลาสเตอร์ ปิดปวดเลือดเย็น ใหญ่ 10x14'),
    (r'พลาสเตอร์ตราเสือ.*?สูตรเย็น.*?10x7', 'พลาสเตอร์ ปิดปวดเลือดเย็น เล็ก 7x10'),
    (r'พลาสเตอร์ตราเสือ.*?สูตรร้อน', 'พลาสเตอร์ ปิดปวดเลือดแดง'),
    (r'พลาสเตอร์ตราเสือ.*?สูตรเย็น', 'พลาสเตอร์ ปิดปวดเลือดเย็น'),
    # Eagle brand
    (r'eagle brand.*?green oil', 'น้ำมันตรานกอินทรีย์'),
    # Alcohol
    (r'แอลกอฮอล์ศิริบัญชา', 'ศิริบัญชา แอลกอฮอล์'),
    # Wangprom herbal
    (r'สมุนไพรวังพรม.*?เขียว', 'วังพรม ยาม่องเลเคด'),
    (r'สมุนไพรวังพรม.*?เหลือง', 'วังพรม ยาม่องไพร'),
    # Gaoeya
    (r'ยากอเอี๊ยะ', 'กอเอี๊ยะ ตราแมว'),
    # Penicillin V → PEN 5แสน (strip IU unit too)
    (r'penicillin\s+v\s+500000', 'pen 5แสน'),
    (r'\biu\b', ''),              # strip dosage unit IU after above substitution
    # Etoricoxib
    (r'etoricoxib', 'nacoxib'),
    # Mega We Care vitamins
    (r'mega\s+we\s+care\s+nat\s+c', 'mg.nat c'),
    (r'mega\s+we\s+care\s+nat\s+b', 'mg.nat b'),
    (r'(?<!\.)nat\s+c\b', 'mg.nat c'),     # negative lookbehind: don't re-expand "mg.nat c"
    (r'(?<!\.)nat\s+b\b', 'mg.nat b'),
    # Normalise vitamin tablet count: strip mg dosage, convert เม็ด→t for size matching
    (r'(mg\.nat\s+[bc])\s+\d+mg\b', r'\1'),
    (r'(mg\.nat\s+[bc])\s+(\d+)เม็ด\b', r'\1 \2t'),  # e.g. 60เม็ด → 60t
    # Perskindol Classic → yellow gel (SCH uses "เหลือง" for classic formulation)
    (r'perskindol\s+classic', 'perskindol เหลือง'),
    (r'mega\s+fish\s+oil', 'mg.fish oil'),
    # Vistra fish oil
    (r'fish\s+oil.*?vistra', 'vistra salmon'),
    # Cetirizine
    (r'\bcetirizine\b', 'cetrizin'),
    # AMK 625 — strip parenthetical active ingredient
    (r'\bamk\s+625\s*\([^)]+\)', 'amk 625'),
    # Fitne
    (r"fitne'?.*?herbal.*?infusion", 'fitne'),
    # Proctosedyl suppositories
    (r'proctosedyl\s+suppositories', 'proctosedyl เหน็บ'),
    # General: strip parenthetical after brand name
    (r'\s*\([^)]{5,}\)', ''),
    # ml/cc equivalence for size tokens
    (r'\b(\d+)\s*ml\b', r'\1cc'),
]

def _preprocess(name: str) -> str:
    s = name.lower().strip()
    for pat, repl in _SCH_SUBS:
        s = re.sub(pat, repl, s, flags=re.IGNORECASE)
    return s.strip()


# ── Conflict pairs: these must NOT match each other ──────────────────────────
# Format: (pattern_in_a, pattern_in_b) — symmetric check
_CONFLICT_PAIRS = [
    (r'nat\s*b', r'nat\s*c'),   # NAT-B vitamin ≠ NAT-C vitamin
    (r'mg\.nat\s*b', r'mg\.nat\s*c'),
]

def _is_conflict(a: str, b: str) -> bool:
    al, bl = a.lower(), b.lower()
    for pa, pb in _CONFLICT_PAIRS:
        if (re.search(pa, al) and re.search(pb, bl)) or \
           (re.search(pb, al) and re.search(pa, bl)):
            return True
    return False


# ── Match order item to delivery items ───────────────────────────────────────

def match_to_delivery(order_name: str, delivery_items: list[dict],
                      threshold: float = 0.65) -> tuple[float, dict | None]:
    best_score, best_item = 0.0, None
    pre_order = _preprocess(order_name)
    for d in delivery_items:
        if _is_conflict(order_name, d["name"]) or _is_conflict(pre_order, d["name"]):
            continue
        # Try both original and preprocessed names
        s = max(
            token_score(order_name, d["name"]),
            token_score(pre_order, d["name"]),
            token_score(pre_order, _preprocess(d["name"])),
        )
        if s > best_score:
            best_score, best_item = s, d
    if best_score >= threshold:
        return best_score, best_item
    return best_score, None


# ── Main comparison ───────────────────────────────────────────────────────────

def compare(order_items: list[tuple[str,int]], delivery: dict) -> dict:
    """
    Returns:
      received    — ordered & delivered (qty match or close)
      diff        — ordered & delivered but qty/size differs
      not_received — ordered but NOT in delivery
      extra       — in delivery but NOT in order

    Uses greedy best-first assignment: highest-scoring (order, delivery) pairs are
    assigned first so that e.g. "BEPANTHEN FIRST AID" claims its exact delivery match
    before "BEPANTHEN OINTMENT" can steal it.
    """
    delivery_items = delivery["items"]

    # Phase 1: score every (order_idx, delivery_item) pair
    candidates: list[tuple[float, int, int, dict]] = []  # (score, order_idx, seq, d_item)
    for i, (drug, _) in enumerate(order_items):
        pre = _preprocess(drug)
        for d in delivery_items:
            if _is_conflict(drug, d["name"]) or _is_conflict(pre, d["name"]):
                continue
            s = max(
                token_score(drug, d["name"]),
                token_score(pre, d["name"]),
                token_score(pre, _preprocess(d["name"])),
            )
            if s >= 0.72:
                candidates.append((s, i, d["seq"], d))

    # Phase 2: greedy assignment — highest scores first
    candidates.sort(key=lambda x: -x[0])
    matched_orders: set[int] = set()
    matched_delivery_seqs: set[int] = set()
    assignments: dict[int, tuple[float, dict]] = {}  # order_idx → (score, d_item)

    for (score, order_idx, deliv_seq, d) in candidates:
        if order_idx not in matched_orders and deliv_seq not in matched_delivery_seqs:
            matched_orders.add(order_idx)
            matched_delivery_seqs.add(deliv_seq)
            assignments[order_idx] = (score, d)

    received    = []
    diff        = []
    not_received = []

    for i, (drug, order_qty) in enumerate(order_items):
        if i not in assignments:
            not_received.append({"drug": drug, "order_qty": order_qty, "best_score": 0.0})
            continue

        score, deliv = assignments[i]
        matched_delivery_seqs.add(deliv["seq"])
        deliv_qty = deliv["qty"]
        notes = []

        # Check qty difference
        qty_match = (order_qty == 0 or order_qty == deliv_qty)

        # Check if name differs significantly (substitution)
        name_diff = score < 0.90

        if qty_match and not name_diff:
            received.append({
                "drug": drug, "order_qty": order_qty,
                "sch_code": deliv["sch_code"], "delivered_name": deliv["name"],
                "delivered_qty": deliv_qty, "unit": deliv["unit"],
                "unit_price": deliv.get("unit_price", 0),
                "score": round(score, 2),
            })
        else:
            if not qty_match and order_qty > 0:
                notes.append(f"สั่ง {order_qty} ส่ง {deliv_qty}")
            if name_diff:
                notes.append(f"ชื่อต่าง (score {score*100:.0f}%)")
            diff.append({
                "drug": drug, "order_qty": order_qty,
                "sch_code": deliv["sch_code"], "delivered_name": deliv["name"],
                "delivered_qty": deliv_qty, "unit": deliv["unit"],
                "unit_price": deliv.get("unit_price", 0),
                "score": round(score, 2),
                "note": " | ".join(notes),
            })

    # Extra items: in delivery but not matched to any order item
    extra = [d for d in delivery_items if d["seq"] not in matched_delivery_seqs]

    return {"received": received, "diff": diff,
            "not_received": not_received, "extra": extra}


# ── Print summary ─────────────────────────────────────────────────────────────

def print_summary(result: dict, order_count: int):
    r, d, n, e = result["received"], result["diff"], result["not_received"], result["extra"]
    total = r + d + n
    W = 100
    print()
    print("=" * W)
    print(f"  เปรียบเทียบ order vs ใบส่งสินค้า")
    print("=" * W)
    print(f"  ออเดอร์รวม: {order_count} รายการ | ✅ ตรงกัน: {len(r)} | ⚠️ ต่างกัน: {len(d)} | ❌ ไม่ได้รับ: {len(n)} | ➕ ส่งเกิน/พิเศษ: {len(e)}")
    print("=" * W)

    if r:
        print(f"\n✅  ได้รับครบ ({len(r)} รายการ)")
        print(f"  {'ชื่อที่สั่ง':<35} {'ชื่อในใบส่ง':<35} {'จำนวน':<8} {'รหัส SCH'}")
        print("  " + "-" * 90)
        for x in sorted(r, key=lambda z: z["drug"]):
            print(f"  {x['drug'][:34]:<35} {x['delivered_name'][:34]:<35} ×{x['delivered_qty']:<7} {x['sch_code']}")

    if d:
        print(f"\n⚠️   ได้รับแต่ต่างกัน ({len(d)} รายการ)")
        print(f"  {'ชื่อที่สั่ง':<32} {'ชื่อในใบส่ง':<32} {'สั่ง':<6} {'ได้':<6} หมายเหตุ")
        print("  " + "-" * 96)
        for x in sorted(d, key=lambda z: z["drug"]):
            print(f"  {x['drug'][:31]:<32} {x['delivered_name'][:31]:<32} ×{str(x['order_qty']):<5} ×{str(x['delivered_qty']):<5} {x['note']}")

    if n:
        print(f"\n❌  ไม่ได้รับ / ไม่พบในใบส่ง ({len(n)} รายการ)")
        for x in sorted(n, key=lambda z: z["drug"]):
            qty_str = f" ×{x['order_qty']}" if x["order_qty"] else ""
            print(f"  • {x['drug']}{qty_str}")

    if e:
        print(f"\n➕  ส่งมาโดยไม่ได้สั่ง / ขนาดพิเศษ ({len(e)} รายการ)")
        for x in sorted(e, key=lambda z: z["name"]):
            print(f"  • [{x['sch_code']}] {x['name']}  ×{x['qty']} {x['unit']}")

    print()


# ── Export Excel ─────────────────────────────────────────────────────────────

def export_excel(result: dict, delivery_meta: dict, merges: list[dict]) -> pathlib.Path:
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("⚠️  pip install openpyxl")
        sys.exit(1)

    EXPORTS.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = EXPORTS / f"compare_{ts}.xlsx"

    wb = openpyxl.Workbook()
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    def _style(cell, fill=None, bold=False, color="000000", align="left", wrap=False):
        if fill: cell.fill = fill
        cell.font = Font(bold=bold, color=color, name="Angsana New", size=13)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        cell.border = thin

    BLUE   = PatternFill("solid", fgColor="1F4E79")
    GREEN  = PatternFill("solid", fgColor="E2EFDA")
    YELLOW = PatternFill("solid", fgColor="FFF2CC")
    RED    = PatternFill("solid", fgColor="FCE4D6")
    PURPLE = PatternFill("solid", fgColor="EAD1DC")
    SUBHDR = PatternFill("solid", fgColor="2E75B6")
    GRAY   = PatternFill("solid", fgColor="D9D9D9")

    # ── Sheet 1: Comparison ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "เปรียบเทียบ"

    inv = delivery_meta.get("invoice_number", "")
    inv_date = delivery_meta.get("date", "")

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"เปรียบเทียบออเดอร์ vs ใบส่งสินค้า  |  Invoice: {inv}  |  วันที่: {inv_date}"
    _style(c, BLUE, bold=True, color="FFFFFF", align="center")
    ws.row_dimensions[1].height = 24

    # Dedup summary if any
    row = 2
    if merges:
        ws.merge_cells(f"A{row}:H{row}")
        merge_strs = ["{} (×{}→×{})".format(m["drug"][:25],
                      "+".join(str(q) for q in m["orig_qtys"]), m["merged_qty"])
                      for m in merges]
        c = ws.cell(row=row, column=1,
                    value="🔁 รายการซ้ำที่ถูกรวม: " + " | ".join(merge_strs))
        _style(c, YELLOW, align="left", wrap=True)
        ws.row_dimensions[row].height = 30
        row += 1

    r, d, n, e = result["received"], result["diff"], result["not_received"], result["extra"]

    def _section(title: str, fill, items: list, cols: list[str], col_w: list[int],
                 row_fn):
        nonlocal row
        ws.merge_cells(f"A{row}:H{row}")
        c = ws.cell(row=row, column=1, value=title)
        _style(c, SUBHDR, bold=True, color="FFFFFF", align="left")
        ws.row_dimensions[row].height = 20
        row += 1

        # Header
        for ci, (h, w) in enumerate(zip(cols, col_w), 1):
            c = ws.cell(row=row, column=ci, value=h)
            _style(c, GRAY, bold=True, align="center")
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[row].height = 20
        row += 1

        # Data
        for item in sorted(items, key=lambda z: z.get("drug", z.get("name", ""))):
            vals = row_fn(item)
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                _style(c, fill, align="center" if ci in (4,5,6) else "left")
            ws.row_dimensions[row].height = 18
            row += 1
        row += 1  # blank separator

    _section(
        f"✅  ได้รับครบ ({len(r)} รายการ)", GREEN,
        r,
        ["ชื่อที่สั่ง","ชื่อในใบส่ง","รหัส SCH","สั่ง","ได้รับ","หน่วย","ราคา/หน่วย","รวม"],
        [30, 30, 10, 7, 7, 8, 12, 12],
        lambda x: [x["drug"], x["delivered_name"], x["sch_code"],
                   x["order_qty"], x["delivered_qty"], x["unit"],
                   x.get("unit_price",""),
                   x.get("unit_price",0) * x["delivered_qty"]],
    )

    _section(
        f"⚠️   ได้รับแต่ต่างกัน ({len(d)} รายการ)", YELLOW,
        d,
        ["ชื่อที่สั่ง","ชื่อในใบส่ง","รหัส SCH","สั่ง","ได้รับ","หน่วย","ราคา/หน่วย","หมายเหตุ"],
        [30, 30, 10, 7, 7, 8, 12, 25],
        lambda x: [x["drug"], x["delivered_name"], x["sch_code"],
                   x["order_qty"], x["delivered_qty"], x["unit"],
                   x.get("unit_price",""), x.get("note","")],
    )

    _section(
        f"❌  ไม่ได้รับ / ไม่พบในใบส่ง ({len(n)} รายการ)", RED,
        n,
        ["ชื่อที่สั่ง","จำนวนที่สั่ง","","","","","","หมายเหตุ"],
        [40, 12, 1, 1, 1, 1, 1, 30],
        lambda x: [x["drug"], x["order_qty"],"","","","","","→ ต้องหาแหล่งซื้ออื่น หรือสั่งซ้ำ"],
    )

    _section(
        f"➕  ส่งมาเพิ่ม / ขนาดพิเศษ ({len(e)} รายการ)", PURPLE,
        e,
        ["ชื่อในใบส่ง","รหัส SCH","จำนวน","หน่วย","ราคา/หน่วย","รวม","","หมายเหตุ"],
        [35, 10, 7, 8, 12, 12, 1, 25],
        lambda x: [x["name"], x["sch_code"], x["qty"], x["unit"],
                   x.get("unit_price",""),
                   x.get("unit_price",0) * x["qty"],
                   "", x.get("note","")],
    )

    ws.freeze_panes = "A3"

    # ── Sheet 2: Full delivery (all 72 items with price) ────────────────────
    ws2 = wb.create_sheet("ใบส่งสินค้าเต็ม")
    HDRS = ["ลำดับ","รหัส SCH","ชื่อสินค้า","จำนวน","หน่วย","ราคา/หน่วย","รวม","หมายเหตุ"]
    CW   = [6, 10, 40, 7, 8, 13, 13, 20]

    ws2.merge_cells("A1:H1")
    c = ws2["A1"]
    c.value = f"ใบส่งสินค้า SCH  |  Invoice: {inv}  |  {inv_date}  |  {delivery_meta.get('customer_name','')}"
    _style(c, BLUE, bold=True, color="FFFFFF", align="center")
    ws2.row_dimensions[1].height = 22

    for ci, (h, w) in enumerate(zip(HDRS, CW), 1):
        c = ws2.cell(row=2, column=ci, value=h)
        _style(c, GRAY, bold=True, align="center")
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[2].height = 20

    for ri, item in enumerate(delivery_meta["items"], start=3):
        vals = [item["seq"], item["sch_code"], item["name"],
                item["qty"], item["unit"], item.get("unit_price",""),
                item.get("total",""), item.get("note","")]
        for ci, v in enumerate(vals, 1):
            c = ws2.cell(row=ri, column=ci, value=v)
            _style(c, align="center" if ci in (1,2,4,5,6,7) else "left")
        ws2.row_dimensions[ri].height = 18

    # Total row
    last = len(delivery_meta["items"]) + 3
    ws2.merge_cells(f"A{last}:F{last}")
    c = ws2.cell(row=last, column=1, value="ยอดรวมทั้งสิ้น")
    _style(c, BLUE, bold=True, color="FFFFFF", align="right")
    c2 = ws2.cell(row=last, column=7, value=delivery_meta.get("total_handwritten",
                  delivery_meta.get("total_printed","")))
    _style(c2, BLUE, bold=True, color="FFFFFF", align="center")
    ws2.row_dimensions[last].height = 22
    ws2.freeze_panes = "A3"

    wb.save(path)
    return path


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Compare order list vs delivery invoice")
    ap.add_argument("--delivery", required=True, help="Path to delivery JSON file")
    ap.add_argument("--export",   action="store_true", help="Export comparison Excel")
    ap.add_argument("--json",     action="store_true", dest="as_json",
                    help="Emit comparison as JSON (for render-html delivery surface)")
    ap.add_argument("file",       nargs="?",    help="Order list file (default: stdin)")
    args = ap.parse_args()

    src   = open(args.file, encoding="utf-8") if args.file else sys.stdin
    lines = [l.strip() for l in src if l.strip()]

    with open(args.delivery, encoding="utf-8") as f:
        delivery = json.load(f)

    order_items, merges = dedup_items(lines)

    result = compare(order_items, delivery)

    if args.as_json:
        # Clean JSON only — no text/merge noise on stdout so it can pipe into render.py
        payload = dict(result)
        payload["merges"] = merges
        payload["order_count"] = len(order_items)
        print(json.dumps(payload, ensure_ascii=False))
        sys.exit(0)

    if merges:
        print(f"\n🔁  รวมรายการซ้ำ {len(merges)} กลุ่ม:")
        for m in merges:
            print("    {}  ×{} → ×{}".format(m["drug"][:40], "+".join(str(q) for q in m["orig_qtys"]), m["merged_qty"]))

    print_summary(result, len(order_items))

    if args.export:
        path = export_excel(result, delivery, merges)
        print(f"📊 Excel saved: {path}")
