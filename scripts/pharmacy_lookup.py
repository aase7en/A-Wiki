#!/usr/bin/env python3
"""
Pharmacy drug order lookup — fuzzy match + SQLite FTS5 + export + history.

Usage:
  echo "amoxicillin\nbetadine 15ml" | python3 scripts/pharmacy_lookup.py
  echo "AMK 650mg: 3\nbetadine: 6"  | python3 scripts/pharmacy_lookup.py --line
  python3 scripts/pharmacy_lookup.py --json          # JSON for Claude
  python3 scripts/pharmacy_lookup.py --export csv    # save CSV
  python3 scripts/pharmacy_lookup.py --export excel  # save Excel (openpyxl)
  python3 scripts/pharmacy_lookup.py --line          # LINE v3 copyable output
  python3 scripts/pharmacy_lookup.py --save          # append to order-history.json
  python3 scripts/pharmacy_lookup.py --export excel --save --line  # all at once

Input format with quantity:
  "drug name: qty"   →  betadine 15ml: 3
  "drug name × qty"  →  betadine 15ml × 3
  "drug name x qty"  →  betadine 15ml x 3
  (no qty = 0, won't appear in LINE format)
"""
import sqlite3, json, re, sys, csv, pathlib, argparse
from datetime import datetime
from difflib import SequenceMatcher

ROOT = pathlib.Path(__file__).parent.parent
sys.path.insert(0, str(ROOT / "scripts"))
from drive_path import get_pharmacy_dir, DriveNotLinkedError  # noqa: E402

DB_PATH    = ROOT / "wiki/entities/pharmacy/drugs.db"
ALIASES_MD = ROOT / "wiki/concepts/pharmacy/drug-aliases.md"


def _pharmacy_path(filename: str) -> pathlib.Path:
    """Resolve a real pharmacy business-data file under drive/pharmacy/.

    Real business records (order history, delivery invoices, exports) live
    only in drive/pharmacy/ — never in the tracked wiki/ tree. Raises
    DriveNotLinkedError with setup instructions if drive/ isn't linked; callers
    that mutate business data should let this propagate and exit(1) rather
    than silently writing into an unsynced fallback directory.
    """
    return get_pharmacy_dir() / filename


def _require_pharmacy_path(filename: str, action: str) -> pathlib.Path:
    """Like _pharmacy_path(), but prints a clear error + exits(1) on failure
    instead of raising — for CLI entry points (--save/--export)."""
    try:
        return _pharmacy_path(filename)
    except DriveNotLinkedError as e:
        print(f"⚠️  Cannot {action}: {e}", file=sys.stderr)
        sys.exit(1)

SHOP_NAME = "ร้านภูฟาร์มาซี"
SHOP_ADDR = "บางปูใหม่ จ.สมุทรปราการ"

CONF_HIGH   = 0.80
CONF_MEDIUM = 0.60

# Qty patterns: "drug: 3" or "drug × 3" or "drug x 3"
_QTY_COLON = re.compile(r":\s*(\d+)\s*$")
_QTY_TIMES = re.compile(r"\s*[×xX]\s*(\d+)\s*$")


# ── Text normalisation ───────────────────────────────────────────────────────

def norm(s: str) -> str:
    return re.sub(r"[\s.\-_()/]+", " ", s.lower()).strip()


# ── Thai→English dosage-form + brand name mapping ────────────────────────────
_THAI_EN: list[tuple[str, str]] = [
    ("เบตาดีน",               "betadine"),
    ("เบต้าดีน",              "betadine"),
    ("ไฮโดรเจนเปอร์ออกไซด์", "hydrogen peroxide"),
    ("ไฮโดรเจน เปอร์ออกไซด์","hydrogen peroxide"),
    ("แอสไพริน",              "aspirin"),
    ("พาราเซตามอล",           "paracetamol"),
    ("ไอบูโพรเฟน",           "ibuprofen"),
    ("ไอบรูเฟน",              "ibuprofen"),
    ("อะม็อกซิซิลลิน",       "amoxicillin"),
    ("คลอร์เฟนิรามีน",        "chlorpheniramine"),
    ("เซทิริซีน",             "cetirizine"),
    ("ซีทิริซีน",             "cetirizine"),
    ("โลราทาดีน",            "loratadine"),
    ("ออมีพราโซล",           "omeprazole"),
    ("เมทฟอร์มิน",           "metformin"),
    ("ไดโคลฟีแนค",          "diclofenac"),
    ("ดิออร์รา",              "diora"),
    ("ดิออร์",                "diora"),
    ("วิตามินซี",             "vitamin c"),
    ("ไรโนแทป",              "rhinotapp"),
    ("ดูลโคแลค",             "dulcolax"),
    ("เวนโทลิน",             "ventolin"),
    ("เซนทรัม",              "centrum"),
    ("โรห์โต",                "rohto"),
    ("นีโอทิเอ",              "neotica"),
    ("ฟูโรเซไมด์",           "furosemide"),
    ("อัลเบนดาโซล",          "albendazole"),
    # Dosage-form terms
    ("ฉีด",                  "injection"),
    ("เม็ด",                 "tablet"),
    ("แคปซูล",               "capsule"),
    ("ครีม",                 "cream"),
    ("โลชั่น",               "lotion"),
    ("เหน็บ",                "suppository"),
    ("ยาน้ำ",                "syrup"),
    ("ยาหยอดตา",             "eye drops"),
    ("ยาหยอดหู",             "ear drops"),
    ("ยาพ่น",                "spray"),
    ("ยาสูด",                "inhaler"),
    ("ยาแผ่น",               "patch"),
    ("ยาอม",                 "lozenge"),
]


def translate_thai(s: str) -> str:
    result = s.lower()
    for th, en in _THAI_EN:
        result = result.replace(th.lower(), en)
    return result


# ── Drug alias dictionary (loaded from drug-aliases.md) ─────────────────────

def load_drug_aliases() -> dict[str, str]:
    """Parse drug-aliases.md tables → {norm(alias): norm(canonical)}."""
    aliases: dict[str, str] = {}
    if not ALIASES_MD.exists():
        return aliases

    header_mode: str | None = None

    for line in ALIASES_MD.read_text(encoding="utf-8").splitlines():
        # Detect table header to know column order
        if "|" not in line:
            header_mode = None
            continue
        if line.startswith("|---") or line.startswith("| ---"):
            continue
        if "ชื่อทางการ" in line or "ชื่อเล่น" in line:
            header_mode = "canonical_first"
            continue
        if "ชื่อที่ส่งมา" in line or "ชื่อที่ถูกต้อง" in line:
            header_mode = "alias_first"
            continue

        if not line.startswith("|"):
            continue
        cols = [c.strip() for c in line.split("|")[1:-1]]
        if len(cols) < 2 or not cols[0] or not cols[1]:
            continue

        if header_mode == "canonical_first":
            # | **CANONICAL** | alias1, alias2 | note |
            canonical = re.sub(r"\*\*(.+?)\*\*", r"\1", cols[0]).strip()
            for alias in cols[1].split(","):
                alias = alias.strip()
                if alias and len(alias) >= 2:
                    aliases[norm(alias)] = norm(canonical)
        elif header_mode == "alias_first":
            # | alias1, alias2 | CANONICAL | cat | note |
            canonical = cols[1].strip()
            for alias in cols[0].split(","):
                alias = alias.strip()
                if alias and len(alias) >= 2:
                    aliases[norm(alias)] = norm(canonical)

    return aliases


# Load aliases once at module import
_DRUG_ALIASES: dict[str, str] = load_drug_aliases()


def expand_query(query: str) -> str:
    """Replace known drug aliases with canonical names."""
    n = norm(query)
    if n in _DRUG_ALIASES:
        return _DRUG_ALIASES[n]
    tokens = n.split()
    changed = False
    for i, t in enumerate(tokens):
        if t in _DRUG_ALIASES:
            tokens[i] = _DRUG_ALIASES[t]
            changed = True
    return " ".join(tokens) if changed else query


def parse_item(line: str) -> tuple[str, int]:
    """Extract (drug_name, qty) from 'drug: qty' or 'drug × qty' formats."""
    m = _QTY_COLON.search(line) or _QTY_TIMES.search(line)
    if m:
        return line[:m.start()].strip(), int(m.group(1))
    return line.strip(), 0


# ── Deduplication ─────────────────────────────────────────────────────────────

def dedup_items(raw_lines: list[str]) -> tuple[list[str], list[dict], list[dict]]:
    """
    Deduplicate input items before lookup.
    - Exact dups (same normalized name): merge, sum qty
    - Near dups (score ≥ 0.85, not identical): flag advisory only
    Returns: (deduped_lines, exact_merge_report, near_dup_report)
    """
    parsed = []
    for line in raw_lines:
        if not line.strip():
            continue
        drug, qty = parse_item(line)
        parsed.append({"drug": drug, "qty": qty, "norm_key": norm(drug), "raw": line})

    # Group by normalized name, preserve insertion order
    groups: dict[str, list] = {}
    key_order: list[str] = []
    for p in parsed:
        k = p["norm_key"]
        if k not in groups:
            key_order.append(k)
        groups.setdefault(k, []).append(p)

    deduped: list[str] = []
    exact_merges: list[dict] = []
    for key in key_order:
        items = groups[key]
        if len(items) == 1:
            deduped.append(items[0]["raw"])
        else:
            total_qty = sum(i["qty"] for i in items)
            best_name = items[0]["drug"]
            merged = f"{best_name}: {total_qty}" if total_qty > 0 else best_name
            deduped.append(merged)
            exact_merges.append({
                "drug":       best_name,
                "count":      len(items),
                "orig_qtys":  [i["qty"] for i in items],
                "merged_qty": total_qty,
            })

    # Near-duplicate detection (same drug, possibly different size/strength)
    near_dups: list[dict] = []
    for i in range(len(key_order)):
        for j in range(i + 1, len(key_order)):
            s = token_score(key_order[i], key_order[j])
            if 0.85 <= s < 1.0:
                near_dups.append({
                    "a":     groups[key_order[i]][0]["drug"],
                    "b":     groups[key_order[j]][0]["drug"],
                    "score": round(s, 2),
                })

    return deduped, exact_merges, near_dups


def _print_dedup_report(merges: list[dict], near_dups: list[dict]):
    if not merges and not near_dups:
        return
    print("\n🔁  ตรวจรายการซ้ำ")
    print("-" * 70)
    if merges:
        print("  รวมรายการซ้ำ (exact):")
        for m in merges:
            qtys = "+".join(str(q) for q in m["orig_qtys"])
            print(f"    ✂️  {m['drug'][:42]:<42}  ×{qtys} → ×{m['merged_qty']}")
    if near_dups:
        print("  อาจซ้ำ — ตรวจสอบก่อน (near-dup ≥85%):")
        for n in near_dups:
            print(f"    ⚠️  {n['a'][:33]:<33}  ≈  {n['b'][:33]:<33}  ({n['score']*100:.0f}%)")
    print()


# ── Scoring ─────────────────────────────────────────────────────────────────

_GENERIC_WORDS = {
    'cream', 'tablet', 'tablets', 'capsule', 'capsules', 'syrup', 'solution',
    'injection', 'drop', 'drops', 'ointment', 'lotion', 'gel', 'spray',
    'powder', 'suspension', 'forte', 'plus', 'oral', 'topical', 'patch',
}


def token_score(a: str, b: str) -> float:
    na, nb = norm(a), norm(b)
    if na == nb:
        return 1.0
    if len(na) > 3 and na in nb:
        return 0.92
    if len(nb) > 3 and nb in na:
        return 0.88
    ta, tb = set(na.split()), set(nb.split())
    overlap = len(ta & tb)
    if overlap == 0:
        return SequenceMatcher(None, na, nb).ratio()
    jaccard = overlap / len(ta | tb)
    if ta and ta <= tb:
        return max(jaccard, 0.85)
    base = max(jaccard, SequenceMatcher(None, na, nb).ratio())
    # Boost if a distinctive brand token (≥4 chars, not a unit/generic word) is shared
    distinctive = {
        t for t in (ta & tb)
        if len(t) >= 4 and not t[:1].isdigit() and t not in _GENERIC_WORDS
    }
    if distinctive:
        return max(base, 0.75)
    return base


def best_score(query: str, row: tuple) -> float:
    candidates = [row[1], row[2] or "", row[3] or ""]
    try:
        candidates.extend(json.loads(row[4] or "[]"))
    except Exception:
        pass
    scores = []
    for c in candidates:
        if not c:
            continue
        scores.append(token_score(query, c))
        tc = translate_thai(c)
        if tc != c.lower():
            scores.append(token_score(query, tc))
    return max(scores) if scores else 0.0


# ── DB search ────────────────────────────────────────────────────────────────

def search_db(conn: sqlite3.Connection, query: str, top_n: int = 3):
    """FTS5 first, fuzzy fallback. Returns [(score, row), ...]"""
    c = conn.cursor()
    results: dict = {}

    # Build all query variants
    translated = translate_thai(query)
    expanded   = expand_query(query)
    all_queries = list(dict.fromkeys([query, translated, expanded]))

    def _score(row) -> float:
        return max(best_score(q, row) for q in all_queries)

    def _run_fts(fts_q: str):
        try:
            rows = c.execute(
                "SELECT d.id,d.name,d.name_th,d.name_en,d.aliases,"
                "d.category_code,d.category_name,d.strength,d.unit,"
                "d.source,d.where_to_buy,d.note "
                "FROM drugs_fts JOIN drugs d ON drugs_fts.rowid=d.id "
                "WHERE drugs_fts MATCH ? ORDER BY rank LIMIT 50",
                (fts_q,),
            ).fetchall()
            for row in rows:
                s = _score(row)
                if row[0] not in results or results[row[0]][0] < s:
                    results[row[0]] = (s, row)
        except sqlite3.OperationalError:
            pass

    # Collect tokens from all query variants
    seen: set[str] = set()
    all_parts: list[str] = []
    for q in all_queries:
        for p in norm(q).split():
            if len(p) >= 2 and p not in seen:
                seen.add(p); all_parts.append(p)

    distinctive = [p for p in all_parts if len(p) >= 4 and not re.match(r"^\d", p)]
    if distinctive:
        _run_fts(" OR ".join(f'"{p}"' for p in distinctive))
    if len(results) < top_n and all_parts:
        _run_fts(" OR ".join(f'"{p}"' for p in all_parts))

    if len(results) < top_n:
        rows = c.execute(
            "SELECT id,name,name_th,name_en,aliases,category_code,category_name,"
            "strength,unit,source,where_to_buy,note FROM drugs "
            "ORDER BY CASE source WHEN 'verified_search' THEN 0 ELSE 1 END LIMIT 5000"
        ).fetchall()
        for row in rows:
            if row[0] in results and results[row[0]][0] >= CONF_HIGH:
                continue
            s = _score(row)
            if s >= CONF_MEDIUM:
                if row[0] not in results or results[row[0]][0] < s:
                    results[row[0]] = (s, row)

    return sorted(results.values(), key=lambda x: -x[0])[:top_n]


def add_to_alt_db(item: dict):
    alt_json = _require_pharmacy_path("alternative-source-items.json", "update verified-search DB")
    data = []
    if alt_json.exists():
        with open(alt_json, encoding="utf-8") as f:
            data = json.load(f)
    existing = {norm(d["name"]) for d in data}
    if norm(item["name"]) not in existing:
        data.append(item)
        with open(alt_json, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    return False


# ── Result builder ───────────────────────────────────────────────────────────
# DB row layout (0-indexed):
# 0=id, 1=name, 2=name_th, 3=name_en, 4=aliases,
# 5=category_code, 6=category_name, 7=strength, 8=unit,
# 9=source, 10=where_to_buy, 11=note

def _build_results(items: list[str]) -> tuple[list, list, list]:
    if not DB_PATH.exists():
        print("⚠️  drugs.db not found → run: python3 scripts/build_pharmacy_db.py")
        sys.exit(1)

    conn = sqlite3.connect(DB_PATH)
    matched, candidates, not_found = [], [], []

    for raw in items:
        raw = raw.strip()
        if not raw:
            continue
        drug, qty = parse_item(raw)
        hits = search_db(conn, drug, top_n=3)

        def _src(row): return "SP" if row[9] == "sp_2020" else "🌐 verified"

        if hits and hits[0][0] >= CONF_HIGH:
            s, row = hits[0]
            matched.append({
                "ordered":       raw,
                "drug":          drug,
                "matched":       row[1],
                "name_th":       row[2] or "",
                "category_code": row[5] or "",
                "category":      row[6] or "อื่นๆ",
                "strength":      row[7] or "",
                "unit":          row[8] or "",
                "source":        _src(row),
                "where_to_buy":  row[10] or "",
                "note":          row[11] or "",
                "confidence":    round(s, 2),
                "qty":           str(qty) if qty else "",
                "status":        "✅",
            })
        elif hits and hits[0][0] >= CONF_MEDIUM:
            s, row = hits[0]
            candidates.append({
                "ordered":       raw,
                "drug":          drug,
                "matched":       row[1],
                "name_th":       row[2] or "",
                "category_code": row[5] or "",
                "category":      row[6] or "อื่นๆ",
                "strength":      row[7] or "",
                "unit":          row[8] or "",
                "source":        _src(row),
                "confidence":    round(s, 2),
                "qty":           str(qty) if qty else "",
                "status":        "⚠️ ต้องยืนยัน",
                "top3":          [(round(sc, 2), r[1]) for sc, r in hits],
                "note":          "",
            })
        else:
            not_found.append({"ordered": raw, "drug": drug, "qty": str(qty) if qty else "", "status": "❌ ไม่พบ"})

    conn.close()
    return matched, candidates, not_found


# ── Print: grouped by category ───────────────────────────────────────────────

def _print_grouped(matched: list, candidates: list, not_found: list):
    W = 110
    print("\n" + "=" * W)
    print(f"📋  สรุปรายการยา — {SHOP_NAME}  ({datetime.now().strftime('%Y-%m-%d %H:%M')})")
    print("=" * W)

    C1, C2, C3, C4, C5 = 28, 36, 10, 8, 12

    def _hdr():
        print(
            f"{'ชื่อที่สั่ง':<{C1}} {'ชื่อยาในฐานข้อมูล':<{C2}}"
            f" {'ความแรง':<{C3}} {'จำนวน':<{C4}} {'แหล่ง':<{C5}} สถานะ"
        )
        print("-" * W)

    def _row(r: dict, suffix: str = ""):
        conf = f"({r['confidence']*100:.0f}%)" if r.get("confidence") else ""
        note = f"  ← {suffix[:45]}" if suffix else ""
        print(
            f"{r['drug'][:C1-1]:<{C1}}"
            f" {r['matched'][:C2-1]:<{C2}}"
            f" {r.get('strength','')[:C3-1]:<{C3}}"
            f" {r.get('qty',''):<{C4}}"
            f" {r.get('source','')[:C4-1]:<{C5}}"
            f" {r['status']} {conf}{note}"
        )

    # ✅ Grouped by category (code + name), A-Z within
    if matched:
        cats: dict[str, list] = {}
        for r in matched:
            code = r.get("category_code", "")
            label = f"{code} {r['category']}".strip() if code else r["category"]
            cats.setdefault(label, []).append(r)
        for cat in sorted(cats):
            print(f"\n📂  {cat}")
            _hdr()
            for r in sorted(cats[cat], key=lambda x: x["matched"].lower()):
                _row(r, r.get("note", ""))

    if candidates:
        print(f"\n⚠️   ต้องยืนยัน {len(candidates)} รายการ")
        _hdr()
        for r in candidates:
            top3_str = " | ".join(f"{n} ({s*100:.0f}%)" for s, n in r["top3"])
            _row(r, f"Top3: {top3_str[:55]}")

    if not_found:
        print(f"\n❌  ไม่พบ {len(not_found)} รายการ → queue subagent search")
        for r in not_found:
            print(f"     • {r['drug']}")

    total = len(matched) + len(candidates) + len(not_found)
    print()
    print(f"รวม: {len(matched)} ✅  {len(candidates)} ⚠️  {len(not_found)} ❌  จากทั้งหมด {total} รายการ")
    print("=" * W + "\n")


# ── LINE v3 format ────────────────────────────────────────────────────────────

def print_line_format(matched: list, candidates: list, not_found: list):
    """
    LINE v3 spec:
    - Show ONLY items with qty > 0
    - No status symbols in output
    - ATC code in category headers: [A02 ยาลดกรด]
    - Format: • name unit × qty
    - Footer: รวม N รายการ (count of items with qty > 0 only)
    """
    items_with_qty = [r for r in matched if int(r.get("qty") or 0) > 0]
    cands_with_qty = [r for r in candidates if int(r.get("qty") or 0) > 0]

    if not items_with_qty and not cands_with_qty:
        print("\n(LINE format: ยังไม่มีรายการที่กรอกจำนวน — ใส่ qty ในรูปแบบ 'ชื่อยา: จำนวน')\n")
        return

    now = datetime.now().strftime("%d/%m/%Y")
    lines = [
        f"📦 รายการสั่งยา — {SHOP_NAME}",
        f"วันที่: {now}",
        "=" * 34,
        "",
    ]

    # Group matched items by ATC code + category name
    cats: dict[str, list] = {}
    for r in items_with_qty:
        code = r.get("category_code", "")
        label = f"{code} {r['category']}".strip() if code else r["category"]
        cats.setdefault(label, []).append(r)

    for cat in sorted(cats):
        lines.append(f"[{cat}]")
        for r in sorted(cats[cat], key=lambda x: x["matched"].lower()):
            name = r["matched"]
            unit = r.get("unit", "")
            unit_str = f" {unit}" if unit else ""
            lines.append(f"• {name}{unit_str} × {r['qty']}")
        lines.append("")

    if cands_with_qty:
        lines.append("[⚠️ ต้องยืนยันก่อนสั่ง]")
        for r in cands_with_qty:
            lines.append(f"• {r['ordered']} → {r['matched']} × {r['qty']}")
        lines.append("")

    total_qty_items = len(items_with_qty) + len(cands_with_qty)
    lines += [
        "=" * 34,
        f"รวม {total_qty_items} รายการ",
    ]
    print("\n" + "\n".join(lines) + "\n")


# ── Export: CSV ──────────────────────────────────────────────────────────────

def export_csv(matched: list, candidates: list, not_found: list) -> pathlib.Path:
    exports_dir = _require_pharmacy_path("exports", "export CSV")
    exports_dir.mkdir(parents=True, exist_ok=True)
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = exports_dir / f"order_{ts}.csv"

    rows, seq = [], 1
    for r in sorted(matched, key=lambda x: (x.get("category_code",""), x["matched"].lower())):
        rows.append([
            seq, r.get("category_code",""), r["category"], r["drug"], r["matched"],
            r["strength"], r["unit"], r["qty"], r["source"],
            f"{r['confidence']*100:.0f}%", "✅", r.get("note",""),
        ])
        seq += 1
    for r in candidates:
        rows.append([seq, r.get("category_code",""), r["category"], r["drug"], r["matched"],
                     r["strength"], r["unit"], r["qty"], r["source"],
                     f"{r['confidence']*100:.0f}%", "⚠️ ต้องยืนยัน", ""])
        seq += 1
    for r in not_found:
        rows.append([seq, "", "", r["drug"], "", "", "", r.get("qty",""), "", "", "❌ ไม่พบ", ""])
        seq += 1

    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["ลำดับ","ATC","หมวดหมู่","ชื่อที่สั่ง","ชื่อยาในฐานข้อมูล",
                    "ความแรง","หน่วย","จำนวน","แหล่ง","ความมั่นใจ","สถานะ","หมายเหตุ"])
        w.writerows(rows)
    return path


# ── Export: Excel ────────────────────────────────────────────────────────────

def export_excel(matched: list, candidates: list, not_found: list) -> pathlib.Path:
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("⚠️  openpyxl not installed → pip install openpyxl")
        sys.exit(1)

    exports_dir = _require_pharmacy_path("exports", "export Excel")
    exports_dir.mkdir(parents=True, exist_ok=True)
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = exports_dir / f"order_{ts}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "รายการสั่งยา"

    BLUE_HDR  = PatternFill("solid", fgColor="1F4E79")
    CAT_FILL  = PatternFill("solid", fgColor="D6E4F0")
    OK_FILL   = PatternFill("solid", fgColor="E2EFDA")
    WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
    ERR_FILL  = PatternFill("solid", fgColor="FCE4D6")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    def _style(cell, fill=None, bold=False, color="000000", align="left"):
        if fill: cell.fill = fill
        cell.font = Font(bold=bold, color=color, name="Angsana New", size=13)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        cell.border = thin

    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value = f"{SHOP_NAME}  —  {SHOP_ADDR}"
    _style(c, BLUE_HDR, bold=True, color="FFFFFF", align="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:L2")
    c = ws["A2"]
    c.value = f"รายการสั่งยา  วันที่: {datetime.now().strftime('%d/%m/%Y  %H:%M น.')}"
    _style(c, PatternFill("solid", fgColor="2E75B6"), bold=True, color="FFFFFF", align="center")
    ws.row_dimensions[2].height = 20

    HEADERS = ["ลำดับ","ATC","หมวดหมู่","ชื่อที่สั่ง","ชื่อยาในฐานข้อมูล",
               "ความแรง","หน่วย","จำนวน","แหล่ง","ความมั่นใจ","สถานะ","หมายเหตุ"]
    COL_W   = [6, 6, 18, 22, 30, 12, 8, 8, 10, 10, 14, 28]

    for col, (h, w) in enumerate(zip(HEADERS, COL_W), start=1):
        c = ws.cell(row=3, column=col, value=h)
        _style(c, BLUE_HDR, bold=True, color="FFFFFF", align="center")
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 22
    ws.freeze_panes = "A4"

    # Flatten all items into one list, sorted by ATC then drug name
    all_rows = []
    for r in matched:
        all_rows.append({**r, "_fill": OK_FILL, "_status": "✅",
                         "_note": r.get("note", "")})
    for r in candidates:
        top3_str = " | ".join(f"{n}({s*100:.0f}%)" for s, n in r.get("top3", []))
        all_rows.append({**r, "_fill": WARN_FILL, "_status": "⚠️ ต้องยืนยัน",
                         "_note": f"Top3: {top3_str}"})
    for r in not_found:
        all_rows.append({**r, "_fill": ERR_FILL, "_status": "❌ ไม่พบ",
                         "_note": ""})

    all_rows.sort(key=lambda x: (x.get("category_code", "ZZ"), x.get("matched", x.get("drug", "")).lower()))

    row_idx = 4
    for seq, r in enumerate(all_rows, start=1):
        top3 = r.get("top3", [])
        note = r.get("_note", r.get("note", ""))
        vals = [
            seq,
            r.get("category_code", ""),
            r.get("category", ""),
            r.get("drug", ""),
            r.get("matched", ""),
            r.get("strength", ""),
            r.get("unit", ""),
            r.get("qty", ""),
            r.get("source", ""),
            f"{r['confidence']*100:.0f}%" if r.get("confidence") else "",
            r.get("_status", ""),
            note,
        ]
        fill = r["_fill"]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=row_idx, column=col, value=v)
            _style(c, fill, align="center" if col in (1, 2, 8, 9, 10) else "left")
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1

    total = len(matched) + len(candidates) + len(not_found)
    ws.merge_cells(f"A{row_idx}:L{row_idx}")
    c = ws.cell(row=row_idx, column=1,
                value=f"รวม: {len(matched)} ✅  {len(candidates)} ⚠️  {len(not_found)} ❌  จากทั้งหมด {total} รายการ")
    _style(c, BLUE_HDR, bold=True, color="FFFFFF", align="center")
    ws.row_dimensions[row_idx].height = 22

    wb.save(path)
    return path


# ── Order history ─────────────────────────────────────────────────────────────

def save_history(items: list[str], matched: list, candidates: list, not_found: list):
    history_json = _require_pharmacy_path("order-history.json", "save order history")
    data: list = []
    if history_json.exists():
        with open(history_json, encoding="utf-8") as f:
            try: data = json.load(f)
            except Exception: data = []

    all_items = (
        [{"ordered": r["drug"], "matched": r["matched"], "category": r["category"],
          "category_code": r.get("category_code",""),
          "strength": r["strength"], "unit": r["unit"], "qty": r["qty"],
          "source": r["source"], "confidence": r["confidence"], "status": r["status"]}
         for r in matched] +
        [{"ordered": r["drug"], "matched": r["matched"], "category": r["category"],
          "category_code": r.get("category_code",""),
          "strength": r["strength"], "unit": r["unit"], "qty": r["qty"],
          "source": r["source"], "confidence": r["confidence"], "status": r["status"]}
         for r in candidates] +
        [{"ordered": r["drug"], "matched": "", "category": "", "category_code": "",
          "strength": "", "unit": "", "qty": r.get("qty",""), "source": "",
          "confidence": 0, "status": "❌"}
         for r in not_found]
    )

    data.append({
        "timestamp":  datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "date":       datetime.now().strftime("%Y-%m-%d"),
        "total":      len(items),
        "matched":    len(matched),
        "candidates": len(candidates),
        "not_found":  len(not_found),
        "items":      all_items,
    })

    with open(history_json, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"📝 บันทึกประวัติแล้ว: {history_json.name}  (รวม {len(data)} session)")


# ── Main lookup ───────────────────────────────────────────────────────────────

def lookup(
    items: list[str],
    json_out: bool = False,
    export: str | None = None,
    line_out: bool = False,
    save_hist: bool = False,
    dedup: bool = True,
) -> dict:
    if dedup:
        items, merges, near = dedup_items(items)
        if merges or near:
            _print_dedup_report(merges, near)

    matched, candidates, not_found = _build_results(items)
    result = {"matched": matched, "candidates": candidates, "not_found": not_found}

    if json_out:
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return result

    _print_grouped(matched, candidates, not_found)

    if line_out:
        print_line_format(matched, candidates, not_found)

    if export == "csv":
        path = export_csv(matched, candidates, not_found)
        print(f"📄 CSV saved: {path}")
    elif export == "excel":
        path = export_excel(matched, candidates, not_found)
        print(f"📊 Excel saved: {path}")

    if save_hist:
        save_history(items, matched, candidates, not_found)

    return result


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Pharmacy drug order lookup")
    ap.add_argument("--json",    action="store_true",  help="Output JSON for Claude")
    ap.add_argument("--export",  choices=["csv","excel"], help="Export to file")
    ap.add_argument("--line",    action="store_true",  help="Print LINE v3 copyable summary")
    ap.add_argument("--save",    action="store_true",  help="Append session to order-history.json")
    ap.add_argument("--no-dedup", action="store_true", help="Skip deduplication step")
    ap.add_argument("file",      nargs="?",            help="Input file (default: stdin)")
    args = ap.parse_args()

    src   = open(args.file, encoding="utf-8") if args.file else sys.stdin
    items = [line.strip() for line in src if line.strip()]

    if not items:
        print("Usage: echo 'betadine 15ml: 3\\nAMK 650mg: 6' | python3 scripts/pharmacy_lookup.py --line")
        sys.exit(1)

    lookup(items, json_out=args.json, export=args.export, line_out=args.line,
           save_hist=args.save, dedup=not args.no_dedup)
